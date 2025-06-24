#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel同步工具配置向导
帮助用户快速配置OneDrive文件路径和工作表名称
"""

import os
import glob
from openpyxl import load_workbook

def find_onedrive_files():
    """查找OneDrive中的Excel文件"""
    onedrive_paths = [
        os.path.expanduser("~/OneDrive"),
        "D:/OneDrive",
        "C:/Users/*/OneDrive",
        "D:/OneDrive - 公司名",  # 企业版OneDrive
    ]
    
    excel_files = []
    for path_pattern in onedrive_paths:
        if "*" in path_pattern:
            # 处理通配符路径
            for user_path in glob.glob(path_pattern):
                excel_path = os.path.join(user_path, "*.xlsx")
                excel_files.extend(glob.glob(excel_path))
        else:
            # 处理固定路径
            if os.path.exists(path_pattern):
                excel_path = os.path.join(path_pattern, "*.xlsx")
                excel_files.extend(glob.glob(excel_path))
    
    return excel_files

def get_sheet_names(file_path):
    """获取Excel文件中的工作表名称"""
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        return sheet_names
    except Exception as e:
        print(f"无法读取文件 {file_path}: {e}")
        return []

def main():
    print("=" * 50)
    print("Excel同步工具配置向导")
    print("=" * 50)
    print()
    
    # 查找OneDrive中的Excel文件
    print("正在搜索OneDrive中的Excel文件...")
    excel_files = find_onedrive_files()
    
    if not excel_files:
        print("未找到OneDrive中的Excel文件")
        print("请手动在config.py中设置SOURCE_FILE_PATH")
        return
    
    print(f"找到 {len(excel_files)} 个Excel文件:")
    for i, file_path in enumerate(excel_files, 1):
        file_name = os.path.basename(file_path)
        print(f"{i}. {file_name}")
        print(f"   路径: {file_path}")
        
        # 显示工作表名称
        sheet_names = get_sheet_names(file_path)
        if sheet_names:
            print(f"   工作表: {', '.join(sheet_names)}")
        print()
    
    # 选择文件
    while True:
        try:
            choice = input("请选择要同步的Excel文件编号 (或输入 'q' 退出): ").strip()
            if choice.lower() == 'q':
                return
            
            choice_num = int(choice)
            if 1 <= choice_num <= len(excel_files):
                selected_file = excel_files[choice_num - 1]
                break
            else:
                print("无效的选择，请重新输入")
        except ValueError:
            print("请输入有效的数字")
    
    # 获取工作表名称
    sheet_names = get_sheet_names(selected_file)
    if not sheet_names:
        print("无法读取工作表名称")
        return
    
    print(f"\n文件 '{os.path.basename(selected_file)}' 包含以下工作表:")
    for i, sheet_name in enumerate(sheet_names, 1):
        print(f"{i}. {sheet_name}")
    
    # 选择要同步的工作表
    print("\n请选择要同步的工作表 (输入编号，用逗号分隔，如: 1,2,3):")
    while True:
        try:
            sheet_choices = input("工作表编号: ").strip()
            choice_indices = [int(x.strip()) - 1 for x in sheet_choices.split(",")]
            
            if all(0 <= i < len(sheet_names) for i in choice_indices):
                selected_sheets = [sheet_names[i] for i in choice_indices]
                break
            else:
                print("无效的选择，请重新输入")
        except ValueError:
            print("请输入有效的数字")
    
    # 生成目标文件夹路径
    base_dir = os.path.dirname(os.path.abspath(__file__))
    target_folders = []
    for sheet_name in selected_sheets:
        # 清理工作表名称作为文件夹名
        folder_name = sheet_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
        target_folder = os.path.join(base_dir, folder_name)
        target_folders.append(target_folder)
    
    # 生成配置文件内容
    config_content = f'''# Excel同步工具配置文件
# 自动生成的配置

# OneDrive中源Excel文件的完整路径
SOURCE_FILE_PATH = r"{selected_file}"

# 要提取的工作表名称（必须与源文件中的工作表名称完全一致）
SHEET_NAMES = {repr(selected_sheets)}

# 目标文件夹路径（提取的工作表将保存到这些文件夹中）
TARGET_FOLDERS = {repr(target_folders)}

# 同步设置
SYNC_SETTINGS = {{
    "monitor_interval": 1,  # 监控间隔（秒）
    "debounce_time": 2,     # 防抖时间（秒）
    "wait_after_save": 1,   # 保存后等待时间（秒）
}}

# 日志设置
LOG_SETTINGS = {{
    "log_file": "excel_sync.log",
    "log_level": "INFO",  # DEBUG, INFO, WARNING, ERROR
    "console_output": True,
}}

# Excel应用程序设置
EXCEL_SETTINGS = {{
    "visible": False,        # Excel应用程序是否可见
    "display_alerts": False, # 是否显示Excel警告
    "enable_events": True,   # 是否启用Excel事件
}}
'''
    
    # 保存配置文件
    config_file = "config.py"
    with open(config_file, "w", encoding="utf-8") as f:
        f.write(config_content)
    
    print(f"\n配置已保存到 {config_file}")
    print("\n配置摘要:")
    print(f"源文件: {selected_file}")
    print(f"工作表: {', '.join(selected_sheets)}")
    print(f"目标文件夹:")
    for folder in target_folders:
        print(f"  - {folder}")
    
    print(f"\n现在可以运行 'python excel_sync_tool.py' 或双击 'start_sync.bat' 来启动同步")

if __name__ == "__main__":
    main() 